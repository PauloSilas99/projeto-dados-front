from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


def _norm(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().strip().lower()


def sanitize_excel_for_engine(src_path: Path) -> Path:
    wb = load_workbook(filename=str(src_path), data_only=True)
    razao_val: Optional[str] = None
    fantasia_cell = None

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                n = _norm(val)
                if n == "razao social":
                    c = ws.cell(row=cell.row, column=cell.column + 1)
                    if c.value and isinstance(c.value, str):
                        razao_val = c.value.strip()
                elif n == "nome fantasia":
                    c = ws.cell(row=cell.row, column=cell.column + 1)
                    fantasia_cell = c

    if fantasia_cell and (fantasia_cell.value is None or str(fantasia_cell.value).strip() == "") and razao_val:
        fantasia_cell.value = razao_val

    out_path = src_path.parent / f"sanitized_{src_path.name}"
    wb.save(str(out_path))
    return out_path